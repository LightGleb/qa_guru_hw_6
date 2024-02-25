import os
from pypdf import PdfReader
import csv
import zipfile
from openpyxl import load_workbook

from paths import RESOURCE_DIR


def test_read_pdf_file_in_archive():
    with (zipfile.ZipFile(os.path.join(RESOURCE_DIR, "archive.zip")) as zip_file):  # открываем архив
        with zip_file.open('Software Testing - Base Course (Svyatoslav Kulikov) - 3rd edition - RU.pdf'
                           ) as pdf_file:  # открываем файл в архиве
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            print(text)
            print(len(reader.pages))

            assert len(reader.pages) == 303


def test_read_xlsx_file_in_archive():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, "archive.zip")) as zip_file:  # открываем архив
        with zip_file.open('Timesheet.xlsx') as xlsx_file:  # открываем файл в архиве
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            print(sheet.cell(row=1, column=2).value)

            assert sheet.cell(row=14, column=6).value == 'Paid'


def test_read_csv_file_in_archive():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, "archive.zip")) as zip_file:  # открываем архив
        with zip_file.open('PUBLIC – Список ценностей в жизни - Ценности.csv') as csv_file:  # открываем файл в архиве
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]
            print(csvreader[1])

            assert second_row[0] == 'Активность'

